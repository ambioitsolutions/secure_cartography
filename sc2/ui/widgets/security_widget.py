#!/usr/bin/env python3
"""
Secure Cartography - Security Widget
=====================================

PyQt6 widget for CVE vulnerability analysis of discovered network devices.
Reads CSV export from Secure Cartography, maps platforms to CPE format,
and queries NIST NVD for known vulnerabilities.

Components split into sc2.ui.widgets.security package for testability.

Standalone testing:
    python security_widget.py

Integration:
    from .widgets.security_widget import SecurityWidget
    widget = SecurityWidget(theme_manager=theme_manager)
"""

import sys
import csv
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional

from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QTableView, QHeaderView, QProgressBar, QPushButton, QLabel,
    QFileDialog, QLineEdit, QTextEdit, QGroupBox,
    QMessageBox, QTabWidget, QDialog, QCheckBox,
)
from PyQt6.QtCore import Qt, QSettings
from PyQt6.QtGui import QPalette, QColor

from .security import (
    ParsedPlatform, PlatformParser, CVECache,
    PlatformTableModel, CVETableModel, CachedVersionsModel, SeverityDelegate,
    SyncWorker,
    CVEDetailDialog, SecurityHelpDialog, AddPatternDialog,
)

logger = logging.getLogger(__name__)


class SecurityWidget(QWidget):
    """
    Main security analysis widget for Secure Cartography

    Features:
    - Load CSV export from discovery
    - Parse platforms and map to CPE format
    - Edit CPE mappings for unrecognized platforms
    - Sync with NIST NVD for vulnerability data
    - View CVE details by version
    - Load and view cached data without new CSV
    - Export vulnerability reports
    - Theme integration with SC
    """

    def __init__(self, db_path: Optional[Path] = None, theme_manager=None, parent=None):
        super().__init__(parent)

        self.db_path = db_path or Path.home() / ".scng" / "cve_cache.db"
        self.cache = CVECache(self.db_path)
        self.parser = PlatformParser()
        self.worker: Optional[SyncWorker] = None
        self.current_csv_path: Optional[str] = None
        self.theme_manager = theme_manager
        self._loaded_platforms = []  # Track loaded platforms for export

        self.settings = QSettings("SecureCartography", "Security")

        self._init_ui()

        # Apply theme if provided
        if theme_manager:
            self.apply_theme(theme_manager.theme)

        # Load cached data on startup
        self._load_from_cache()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # Top controls - Row 1
        top_bar = QHBoxLayout()

        self.load_btn = QPushButton("Load CSV")
        self.load_btn.clicked.connect(self._load_csv)
        top_bar.addWidget(self.load_btn)

        self.load_cache_btn = QPushButton("View Cache")
        self.load_cache_btn.setToolTip("View previously synced versions from cache")
        self.load_cache_btn.clicked.connect(self._load_from_cache)
        top_bar.addWidget(self.load_cache_btn)

        top_bar.addSpacing(20)

        top_bar.addWidget(QLabel("API Key:"))
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("Optional - increases rate limit")
        self.api_key_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.api_key_input.setText(self.settings.value("nvd_api_key", ""))
        self.api_key_input.setMaximumWidth(200)
        top_bar.addWidget(self.api_key_input)

        top_bar.addStretch()

        # Help button
        self.help_btn = QPushButton("? Help")
        self.help_btn.setToolTip("Show help for security analysis workflow")
        self.help_btn.clicked.connect(self._show_help)
        top_bar.addWidget(self.help_btn)

        top_bar.addSpacing(10)

        # Force re-sync checkbox
        self.force_sync_cb = QCheckBox("Force Re-sync")
        self.force_sync_cb.setToolTip("Re-sync versions even if already in cache")
        top_bar.addWidget(self.force_sync_cb)

        self.sync_btn = QPushButton("Sync Selected")
        self.sync_btn.clicked.connect(self._start_sync)
        self.sync_btn.setEnabled(False)
        top_bar.addWidget(self.sync_btn)

        self.stop_btn = QPushButton("Stop")
        self.stop_btn.clicked.connect(self._stop_sync)
        self.stop_btn.setEnabled(False)
        top_bar.addWidget(self.stop_btn)

        layout.addLayout(top_bar)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        self.status_label = QLabel("Load a CSV export from Secure Cartography to begin")
        layout.addWidget(self.status_label)

        # Main content splitter
        splitter = QSplitter(Qt.Orientation.Vertical)

        # Platform table
        platform_group = QGroupBox("Discovered Platforms")
        platform_layout = QVBoxLayout(platform_group)

        self.platform_model = PlatformTableModel()
        self.platform_table = QTableView()
        self.platform_table.setModel(self.platform_model)
        self.platform_table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.platform_table.setAlternatingRowColors(True)
        self.platform_table.horizontalHeader().setStretchLastSection(True)
        self.platform_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.platform_table.selectionModel().selectionChanged.connect(self._on_platform_selected)

        platform_layout.addWidget(self.platform_table)

        # Platform action buttons
        platform_actions = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All Valid")
        self.select_all_btn.clicked.connect(self._select_all_valid)
        platform_actions.addWidget(self.select_all_btn)

        self.add_pattern_btn = QPushButton("Add Pattern...")
        self.add_pattern_btn.clicked.connect(self._add_custom_pattern)
        platform_actions.addWidget(self.add_pattern_btn)

        platform_actions.addStretch()

        self.platform_count_label = QLabel("")
        platform_actions.addWidget(self.platform_count_label)

        platform_layout.addLayout(platform_actions)
        splitter.addWidget(platform_group)

        # CVE results (tabs)
        results_group = QGroupBox("CVE Results")
        results_layout = QVBoxLayout(results_group)

        self.results_tabs = QTabWidget()

        # Cached versions tab (from database)
        cached_widget = QWidget()
        cached_layout = QVBoxLayout(cached_widget)

        cached_header = QHBoxLayout()
        self.cached_count_label = QLabel("0 versions in cache")
        cached_header.addWidget(self.cached_count_label)
        cached_header.addStretch()

        refresh_cache_btn = QPushButton("Refresh")
        refresh_cache_btn.clicked.connect(self._load_from_cache)
        cached_header.addWidget(refresh_cache_btn)

        cached_layout.addLayout(cached_header)

        self.cached_model = CachedVersionsModel()
        self.cached_table = QTableView()
        self.cached_table.setModel(self.cached_model)
        self.cached_table.setAlternatingRowColors(True)
        self.cached_table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.cached_table.horizontalHeader().setStretchLastSection(True)
        self.cached_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.cached_table.doubleClicked.connect(self._on_cached_version_double_click)
        cached_layout.addWidget(self.cached_table)
        self.results_tabs.addTab(cached_widget, "Cached Versions")

        # Summary tab
        summary_widget = QWidget()
        summary_layout = QVBoxLayout(summary_widget)
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        summary_layout.addWidget(self.summary_text)
        self.results_tabs.addTab(summary_widget, "Summary")

        # CVE list tab (for selected version)
        cve_widget = QWidget()
        cve_layout = QVBoxLayout(cve_widget)

        self.cve_version_label = QLabel("Select a version to view CVEs")
        cve_layout.addWidget(self.cve_version_label)

        self.cve_model = CVETableModel()
        self.cve_table = QTableView()
        self.cve_table.setModel(self.cve_model)
        self.cve_table.setItemDelegate(SeverityDelegate(self.cve_model, self.cve_table))
        self.cve_table.setAlternatingRowColors(False)  # Delegate handles colors
        self.cve_table.horizontalHeader().setStretchLastSection(True)
        self.cve_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.cve_table.doubleClicked.connect(self._on_cve_double_click)
        cve_layout.addWidget(self.cve_table)
        self.results_tabs.addTab(cve_widget, "CVEs")

        results_layout.addWidget(self.results_tabs)

        # Export buttons (always visible, below tabs)
        export_row = QHBoxLayout()
        export_row.addStretch()

        export_report_btn = QPushButton("Export Full Report")
        export_report_btn.setToolTip("Export all CVEs with affected platforms to Excel")
        export_report_btn.clicked.connect(self._export_full_report)
        export_row.addWidget(export_report_btn)

        export_devices_btn = QPushButton("Export by Device")
        export_devices_btn.setToolTip("Export vulnerability summary per device to Excel")
        export_devices_btn.clicked.connect(self._export_devices_report)
        export_row.addWidget(export_devices_btn)

        results_layout.addLayout(export_row)

        splitter.addWidget(results_group)

        splitter.setSizes([300, 200])
        layout.addWidget(splitter)

        self._update_summary()

    def _load_csv(self):
        """Load CSV export from Secure Cartography"""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Load Discovery CSV", "", "CSV Files (*.csv);;All Files (*)"
        )
        if not filepath:
            return

        self.current_csv_path = filepath

        try:
            # Track platform -> list of device names
            platforms_dict = {}  # platform_string -> {'count': int, 'devices': list}

            with open(filepath, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Look for platform column (flexible naming)
                    platform = None
                    for key in ['platform', 'Platform', 'os', 'OS', 'software', 'version']:
                        if key in row and row[key]:
                            platform = row[key].strip()
                            break

                    # Get device identifier (hostname or IP)
                    device_name = None
                    for key in ['hostname', 'Hostname', 'name', 'Name', 'ip', 'IP',
                                'ip_address', 'IP Address', 'management_ip', 'device']:
                        if key in row and row[key]:
                            device_name = row[key].strip()
                            break

                    if platform:
                        if platform not in platforms_dict:
                            platforms_dict[platform] = {'count': 0, 'devices': []}
                        platforms_dict[platform]['count'] += 1
                        if device_name and device_name not in platforms_dict[platform]['devices']:
                            platforms_dict[platform]['devices'].append(device_name)

            # Parse each unique platform
            parsed = []
            for platform_str, info in platforms_dict.items():
                p = self.parser.parse(platform_str)
                p.device_count = info['count']
                p.device_names = info['devices']
                parsed.append(p)

            # Sort by device count descending
            parsed.sort(key=lambda x: x.device_count, reverse=True)

            self.platform_model.load_platforms(parsed)

            # Store for export
            self._loaded_platforms = parsed

            self.sync_btn.setEnabled(True)

            valid = sum(1 for p in parsed if p.confidence == "high")
            self.platform_count_label.setText(
                f"{len(parsed)} unique platforms ({valid} auto-mapped)"
            )
            total_devices = sum(info['count'] for info in platforms_dict.values())
            self.status_label.setText(f"Loaded {total_devices} devices from {filepath}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load CSV: {e}")
            logger.exception("CSV load error")

    def _load_from_cache(self):
        """Load and display previously synced versions from cache"""
        # Reconnect to ensure we see latest data
        self.cache.reconnect()

        versions = self.cache.get_version_summary()
        self.cached_model.load_versions(versions)
        self.cached_count_label.setText(f"{len(versions)} versions in cache")

        if versions:
            self.results_tabs.setCurrentIndex(0)  # Switch to cached versions tab
            self.status_label.setText(f"Loaded {len(versions)} versions from cache ({self.db_path})")
        else:
            self.status_label.setText("Cache is empty - load a CSV and sync to populate")

        self._update_summary()

    def _on_cached_version_double_click(self, index):
        """Show CVE detail dialog for double-clicked cached version"""
        v = self.cached_model.get_version_at(index.row())
        if not v:
            return

        cves = self.cache.get_cves_for_version(v['vendor'], v['product'], v['version'])

        dialog = CVEDetailDialog(
            v['vendor'], v['product'], v['version'], cves, self
        )
        dialog.exec()

    def _on_cve_double_click(self, index):
        """Open CVE in browser on double-click"""
        row = index.row()
        if row < len(self.cve_model.cves):
            cve_id = self.cve_model.cves[row].get('cve_id', '')
            if cve_id:
                import webbrowser
                webbrowser.open(f"https://nvd.nist.gov/vuln/detail/{cve_id}")

    def _export_full_report(self):
        """Export full vulnerability report to Excel with affected devices"""
        versions = self.cache.get_version_summary()
        if not versions:
            QMessageBox.warning(self, "No Data", "No cached data to export.")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Export Vulnerability Report",
            f"vulnerability_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        # Build lookup from loaded platforms (if available)
        platform_devices = {}
        if hasattr(self, '_loaded_platforms') and self._loaded_platforms:
            for p in self._loaded_platforms:
                key = f"{p.cpe_vendor}:{p.cpe_product}:{p.cpe_version}"
                platform_devices[key] = {
                    'count': p.device_count,
                    'devices': p.device_names
                }

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "CVE Report"

            # Severity colors
            severity_fills = {
                'CRITICAL': PatternFill('solid', fgColor='DC2626'),
                'HIGH': PatternFill('solid', fgColor='EA580C'),
                'MEDIUM': PatternFill('solid', fgColor='CA8A04'),
                'LOW': PatternFill('solid', fgColor='16A34A'),
            }
            white_font = Font(color='FFFFFF', bold=True)
            header_fill = PatternFill('solid', fgColor='1F2937')
            header_font = Font(bold=True, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ['Vendor', 'Product', 'Version', 'Device Count', 'Affected Devices',
                       'CVE ID', 'Severity', 'CVSS v3', 'Published', 'Description']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data rows
            row_num = 2
            for v in versions:
                cves = self.cache.get_cves_for_version(v['vendor'], v['product'], v['version'])

                # Look up device info
                key = f"{v['vendor']}:{v['product']}:{v['version']}"
                device_info = platform_devices.get(key, {'count': 0, 'devices': []})
                device_count = device_info['count'] or ''
                device_list = '; '.join(device_info['devices'])

                if cves:
                    for cve in cves:
                        severity = cve.get('severity', '').upper()

                        ws.cell(row=row_num, column=1, value=v['vendor'])
                        ws.cell(row=row_num, column=2, value=v['product'])
                        ws.cell(row=row_num, column=3, value=v['version'])
                        ws.cell(row=row_num, column=4, value=device_count)
                        ws.cell(row=row_num, column=5, value=device_list)
                        ws.cell(row=row_num, column=6, value=cve.get('cve_id', ''))

                        sev_cell = ws.cell(row=row_num, column=7, value=severity)
                        if severity in severity_fills:
                            sev_cell.fill = severity_fills[severity]
                            sev_cell.font = white_font
                            sev_cell.alignment = Alignment(horizontal='center')

                        ws.cell(row=row_num, column=8, value=cve.get('cvss_v3_score', ''))
                        ws.cell(row=row_num, column=9,
                                value=cve.get('published_date', '')[:10] if cve.get('published_date') else '')

                        desc_cell = ws.cell(row=row_num, column=10, value=cve.get('description', ''))
                        desc_cell.alignment = Alignment(wrap_text=True)

                        for col in range(1, 11):
                            ws.cell(row=row_num, column=col).border = thin_border

                        row_num += 1
                else:
                    ws.cell(row=row_num, column=1, value=v['vendor'])
                    ws.cell(row=row_num, column=2, value=v['product'])
                    ws.cell(row=row_num, column=3, value=v['version'])
                    ws.cell(row=row_num, column=4, value=device_count)
                    ws.cell(row=row_num, column=5, value=device_list)
                    ws.cell(row=row_num, column=10, value='No CVEs found')
                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).border = thin_border
                    row_num += 1

            # Column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 80

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f"A1:J{row_num - 1}"

            wb.save(filepath)

            total_rows = row_num - 2
            QMessageBox.information(
                self, "Export Complete",
                f"Exported {total_rows} rows for {len(versions)} versions to:\n{filepath}"
            )

        except ImportError:
            QMessageBox.critical(self, "Missing Dependency",
                                 "openpyxl is required for Excel export.\nInstall with: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
            logger.exception("Export error")

    def _export_devices_report(self):
        """Export device-centric vulnerability report to Excel"""
        if not hasattr(self, '_loaded_platforms') or not self._loaded_platforms:
            QMessageBox.warning(self, "No Data",
                                "Load a CSV first to export device-focused report.")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Export Devices Vulnerability Report",
            f"devices_vulnerabilities_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "Device Vulnerabilities"

            severity_fills = {
                'critical': PatternFill('solid', fgColor='DC2626'),
                'high': PatternFill('solid', fgColor='EA580C'),
                'medium': PatternFill('solid', fgColor='CA8A04'),
                'low': PatternFill('solid', fgColor='16A34A'),
            }
            white_font = Font(color='FFFFFF', bold=True)
            header_fill = PatternFill('solid', fgColor='1F2937')
            header_font = Font(bold=True, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            headers = ['Device', 'Platform', 'Vendor', 'Product', 'Version',
                       'Total CVEs', 'Critical', 'High', 'Medium', 'Low']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row_num = 2
            for p in self._loaded_platforms:
                if p.cpe_vendor and p.cpe_product and p.cpe_version:
                    cves = self.cache.get_cves_for_version(
                        p.cpe_vendor, p.cpe_product, p.cpe_version
                    )
                    counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
                    for cve in cves:
                        sev = cve.get('severity', '').upper()
                        if sev in counts:
                            counts[sev] += 1
                    total = len(cves)
                else:
                    total = 0
                    counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}

                for device in p.device_names:
                    ws.cell(row=row_num, column=1, value=device)
                    ws.cell(row=row_num, column=2, value=p.raw)
                    ws.cell(row=row_num, column=3, value=p.cpe_vendor or p.vendor)
                    ws.cell(row=row_num, column=4, value=p.cpe_product or p.product)
                    ws.cell(row=row_num, column=5, value=p.cpe_version or p.version)
                    ws.cell(row=row_num, column=6, value=total)

                    crit_cell = ws.cell(row=row_num, column=7, value=counts['CRITICAL'])
                    if counts['CRITICAL'] > 0:
                        crit_cell.fill = severity_fills['critical']
                        crit_cell.font = white_font
                    crit_cell.alignment = Alignment(horizontal='center')

                    high_cell = ws.cell(row=row_num, column=8, value=counts['HIGH'])
                    if counts['HIGH'] > 0:
                        high_cell.fill = severity_fills['high']
                        high_cell.font = white_font
                    high_cell.alignment = Alignment(horizontal='center')

                    med_cell = ws.cell(row=row_num, column=9, value=counts['MEDIUM'])
                    if counts['MEDIUM'] > 0:
                        med_cell.fill = severity_fills['medium']
                        med_cell.font = white_font
                    med_cell.alignment = Alignment(horizontal='center')

                    low_cell = ws.cell(row=row_num, column=10, value=counts['LOW'])
                    if counts['LOW'] > 0:
                        low_cell.fill = severity_fills['low']
                        low_cell.font = white_font
                    low_cell.alignment = Alignment(horizontal='center')

                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).border = thin_border

                    row_num += 1

                if not p.device_names and p.device_count:
                    ws.cell(row=row_num, column=1, value=f"({p.device_count} devices)")
                    ws.cell(row=row_num, column=2, value=p.raw)
                    ws.cell(row=row_num, column=3, value=p.cpe_vendor or p.vendor)
                    ws.cell(row=row_num, column=4, value=p.cpe_product or p.product)
                    ws.cell(row=row_num, column=5, value=p.cpe_version or p.version)
                    ws.cell(row=row_num, column=6, value=total)
                    ws.cell(row=row_num, column=7, value=counts['CRITICAL'])
                    ws.cell(row=row_num, column=8, value=counts['HIGH'])
                    ws.cell(row=row_num, column=9, value=counts['MEDIUM'])
                    ws.cell(row=row_num, column=10, value=counts['LOW'])
                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).border = thin_border
                    row_num += 1

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 10

            ws.freeze_panes = 'A2'
            ws.auto_filter.ref = f"A1:J{row_num - 1}"

            wb.save(filepath)

            total_rows = row_num - 2
            QMessageBox.information(
                self, "Export Complete",
                f"Exported {total_rows} device vulnerability records to:\n{filepath}"
            )

        except ImportError:
            QMessageBox.critical(self, "Missing Dependency",
                                 "openpyxl is required for Excel export.\nInstall with: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
            logger.exception("Export error")

    def _select_all_valid(self):
        """Select all rows with valid CPE mappings"""
        self.platform_table.clearSelection()
        selection = self.platform_table.selectionModel()

        for i, p in enumerate(self.platform_model.platforms):
            if p.cpe_vendor and p.cpe_product and p.cpe_version:
                idx = self.platform_model.index(i, 0)
                selection.select(idx, selection.SelectionFlag.Select | selection.SelectionFlag.Rows)

    def _on_platform_selected(self):
        """Handle platform selection change"""
        rows = set(idx.row() for idx in self.platform_table.selectedIndexes())
        if len(rows) == 1:
            row = list(rows)[0]
            p = self.platform_model.platforms[row]

            # Load CVEs for this version if synced
            if p.cpe_vendor and p.cpe_product and p.cpe_version:
                cves = self.cache.get_cves_for_version(p.cpe_vendor, p.cpe_product, p.cpe_version)
                self.cve_model.load_cves(cves)
                self.cve_version_label.setText(
                    f"CVEs for {p.cpe_vendor}:{p.cpe_product}:{p.cpe_version} ({len(cves)} found)"
                )

                if cves:
                    self.results_tabs.setCurrentIndex(2)  # Switch to CVE tab
            else:
                self.cve_model.load_cves([])
                self.cve_version_label.setText("Invalid CPE mapping - edit vendor/product/version fields")

    def _start_sync(self):
        """Start NVD sync for selected platforms"""
        rows = set(idx.row() for idx in self.platform_table.selectedIndexes())
        platforms = self.platform_model.get_selected_for_sync(list(rows))

        if not platforms:
            QMessageBox.warning(self, "No Valid Selections",
                                "Select platforms with valid CPE mappings to sync.")
            return

        force_sync = self.force_sync_cb.isChecked()

        # Filter out already-synced unless force is checked
        if not force_sync:
            to_sync = []
            skipped = 0
            for p in platforms:
                if not self.cache.is_version_synced(p.cpe_vendor, p.cpe_product, p.cpe_version):
                    to_sync.append(p)
                else:
                    skipped += 1
                    self.platform_model.update_status(p.raw, "synced")

            if skipped > 0:
                self.status_label.setText(f"Skipped {skipped} already-synced versions")

            platforms = to_sync

        if not platforms:
            QMessageBox.information(self, "Nothing to Sync",
                                    "All selected versions are already synced. Check 'Force Re-sync' to update them.")
            return

        # Save API key
        api_key = self.api_key_input.text().strip() or None
        if api_key:
            self.settings.setValue("nvd_api_key", api_key)

        # Update UI
        self.sync_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setVisible(True)
        self.progress.setMaximum(len(platforms))
        self.progress.setValue(0)

        # Mark all as syncing
        for p in platforms:
            self.platform_model.update_status(p.raw, "syncing")

        # Start worker
        self.worker = SyncWorker(self.db_path, platforms, api_key)
        self.worker.progress.connect(self._on_sync_progress)
        self.worker.version_complete.connect(self._on_version_complete)
        self.worker.finished_all.connect(self._on_sync_finished)
        self.worker.start()

    def _stop_sync(self):
        """Stop the sync worker"""
        if self.worker:
            self.worker.stop()

    def _on_sync_progress(self, current: int, total: int, message: str):
        """Handle sync progress update"""
        self.progress.setValue(current)
        self.status_label.setText(message)

    def _on_version_complete(self, raw: str, result: dict):
        """Handle single version sync complete"""
        status = result.get("status", "error")
        self.platform_model.update_status(raw, status)

    def _on_sync_finished(self, summary: dict):
        """Handle sync complete"""
        self.sync_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress.setVisible(False)

        self.status_label.setText(
            f"Sync complete: {summary['synced']} versions, "
            f"{summary['total_cves']} CVEs found, {summary['errors']} errors"
        )

        # Reconnect to see data written by worker thread
        self.cache.reconnect()

        self._update_summary()
        self._load_from_cache()  # Refresh cached versions table
        self.worker = None

    def _update_summary(self):
        """Update the summary text"""
        overall = self.cache.get_overall_summary()
        versions = self.cache.get_version_summary()

        text = "=== CVE Cache Summary ===\n\n"
        text += f"Total CVEs: {overall.get('total', 0)}\n"
        text += f"  Critical: {overall.get('critical', 0)}\n"
        text += f"  High: {overall.get('high', 0)}\n"
        text += f"  Medium: {overall.get('medium', 0)}\n"
        text += f"  Low: {overall.get('low', 0)}\n\n"

        if versions:
            text += "=== Synced Versions ===\n\n"
            for v in versions[:20]:
                text += f"{v['vendor']}:{v['product']}:{v['version']}\n"
                text += f"  CVEs: {v['cve_count']} (C:{v['critical_count']} H:{v['high_count']} M:{v['medium_count']} L:{v['low_count']})\n"
                text += f"  Synced: {v['last_synced'][:10] if v['last_synced'] else 'Never'}\n\n"

        self.summary_text.setText(text)

    def _add_custom_pattern(self):
        """Show dialog to add a custom platform pattern"""
        dialog = AddPatternDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            pattern = dialog.get_pattern()
            self.parser.save_custom_pattern(**pattern)

            # Re-parse current platforms
            if self.platform_model.platforms:
                for p in self.platform_model.platforms:
                    if p.confidence == "low":
                        new_p = self.parser.parse(p.raw)
                        if new_p.confidence == "high":
                            p.cpe_vendor = new_p.cpe_vendor
                            p.cpe_product = new_p.cpe_product
                            p.cpe_version = new_p.cpe_version
                            p.confidence = "high"

                self.platform_model.layoutChanged.emit()

            QMessageBox.information(self, "Pattern Added",
                                    f"Pattern saved to {self.parser.custom_patterns_path}")

    def _show_help(self):
        """Show help dialog"""
        dialog = SecurityHelpDialog(self)
        if self.theme_manager:
            # Apply theme to help dialog
            theme = self.theme_manager.theme
            dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {theme.bg_primary};
                    color: {theme.text_primary};
                }}
                QTextEdit {{
                    background-color: {theme.bg_secondary};
                    color: {theme.text_primary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                }}
                QPushButton {{
                    background-color: {theme.accent};
                    color: {theme.text_on_accent};
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {theme.accent_hover};
                }}
                QLabel {{
                    color: {theme.text_primary};
                }}
            """)
        dialog.exec()

    def apply_theme(self, theme):
        """Apply theme colors to widget"""
        if hasattr(theme, 'bg_primary'):
            self._current_theme = theme

            self.setStyleSheet(f"""
                QWidget {{
                    background-color: {theme.bg_primary};
                    color: {theme.text_primary};
                }}

                QGroupBox {{
                    background-color: {theme.bg_secondary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 8px;
                    margin-top: 12px;
                    padding-top: 8px;
                    font-weight: bold;
                }}

                QGroupBox::title {{
                    subcontrol-origin: margin;
                    left: 12px;
                    padding: 0 8px;
                    color: {theme.text_primary};
                }}

                QPushButton {{
                    background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    padding: 8px 16px;
                    color: {theme.text_primary};
                }}

                QPushButton:hover {{
                    border-color: {theme.accent};
                    color: {theme.accent};
                }}

                QPushButton:pressed {{
                    background-color: {theme.bg_hover};
                }}

                QPushButton:disabled {{
                    background-color: {theme.bg_disabled};
                    color: {theme.text_disabled};
                    border-color: {theme.border_dim};
                }}

                QLineEdit {{
                    background-color: {theme.bg_input};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    padding: 8px;
                    color: {theme.text_primary};
                }}

                QLineEdit:focus {{
                    border-color: {theme.accent};
                }}

                QTableView {{
                    background-color: {theme.bg_secondary};
                    alternate-background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    gridline-color: {theme.border_dim};
                    color: {theme.text_primary};
                }}

                QTableView::item:selected {{
                    background-color: {theme.bg_selected};
                    color: {theme.accent};
                }}

                QHeaderView::section {{
                    background-color: {theme.bg_tertiary};
                    color: {theme.text_primary};
                    border: none;
                    border-bottom: 1px solid {theme.border_dim};
                    padding: 8px;
                    font-weight: bold;
                }}

                QTabWidget::pane {{
                    background-color: {theme.bg_secondary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                }}

                QTabBar::tab {{
                    background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-bottom: none;
                    border-top-left-radius: 6px;
                    border-top-right-radius: 6px;
                    padding: 8px 16px;
                    color: {theme.text_secondary};
                }}

                QTabBar::tab:selected {{
                    background-color: {theme.bg_secondary};
                    color: {theme.accent};
                }}

                QTabBar::tab:hover:!selected {{
                    color: {theme.text_primary};
                }}

                QTextEdit {{
                    background-color: {theme.bg_secondary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    color: {theme.text_primary};
                }}

                QProgressBar {{
                    background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    text-align: center;
                    color: {theme.text_primary};
                }}

                QProgressBar::chunk {{
                    background-color: {theme.accent};
                    border-radius: 5px;
                }}

                QCheckBox {{
                    color: {theme.text_primary};
                    spacing: 8px;
                }}

                QCheckBox::indicator {{
                    width: 18px;
                    height: 18px;
                    border: 1px solid {theme.border_secondary};
                    border-radius: 4px;
                    background-color: {theme.bg_input};
                }}

                QCheckBox::indicator:checked {{
                    background-color: {theme.accent};
                    border-color: {theme.accent};
                }}

                QLabel {{
                    color: {theme.text_primary};
                }}

                QScrollBar:vertical {{
                    background-color: {theme.bg_primary};
                    width: 10px;
                    margin: 0;
                }}

                QScrollBar::handle:vertical {{
                    background-color: {theme.scrollbar_handle};
                    border-radius: 5px;
                    min-height: 20px;
                }}

                QScrollBar::handle:vertical:hover {{
                    background-color: {theme.scrollbar_hover};
                }}

                QScrollBar::add-line:vertical,
                QScrollBar::sub-line:vertical {{
                    height: 0px;
                }}
            """)

    def closeEvent(self, event):
        """Clean up on close"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait()
        self.cache.close()
        super().closeEvent(event)


# Standalone Entry Point
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Dark mode option
    if "--dark" in sys.argv:
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
        palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Base, QColor(25, 25, 25))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
        app.setPalette(palette)

    widget = SecurityWidget()
    widget.setWindowTitle("Secure Cartography - Security Analysis")
    widget.resize(1000, 700)
    widget.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
